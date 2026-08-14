"""Generate the full ASR preprocessing hotword table from addresses and slots."""

from __future__ import annotations

import argparse
from pathlib import Path
import re

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOKS = (
    Path(
        "/home/twai/wjl/AddressBot/data_v4/深圳/"
        "AddressBot_shenzhen_with_real_data_selected_duplicates_merged.xlsx"
    ),
    Path(
        "/home/twai/wjl/AddressBot_v3/data_v4/深圳/"
        "AddressBot_shenzhen_with_real_data_software_park_phase2_aoi3.xlsx"
    ),
)
DEFAULT_OUTPUT_DIR = BASE_DIR / "hotwords_full"
DEFAULT_DYNAMIC_DIR = BASE_DIR / "hotwords"
DEFAULT_SLOTS_FILE = DEFAULT_OUTPUT_DIR / "slots.txt"
DEFAULT_ADDRESS_FILE = DEFAULT_OUTPUT_DIR / "address.txt"
DEFAULT_FULL_FILE = DEFAULT_OUTPUT_DIR / "full.txt"
DYNAMIC_RELATIVE_FILES = (
    Path("address.txt"),
    Path("inquiry_fire_base.txt"),
    Path("scenes/highrise.txt"),
    Path("scenes/crowded_place.txt"),
    Path("scenes/chemical.txt"),
    Path("scenes/elevator.txt"),
)
BRACKET_TRANSLATION = str.maketrans({char: "" for char in "()（）[]【】{}｛｝"})


def normalize_hotword(value) -> str:
    """Remove brackets and whitespace without dropping bracketed content."""
    text = str(value or "").strip().translate(BRACKET_TRANSLATION)
    return re.sub(r"\s+", "", text)


def _unique_terms(values) -> list[str]:
    return list(dict.fromkeys(term for value in values if (term := normalize_hotword(value))))


def load_address_terms(workbook_path: str | Path) -> list[str]:
    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        if "name" not in headers:
            raise ValueError(f"address workbook has no 'name' column: {workbook_path}")
        name_index = headers.index("name")
        return _unique_terms(row[name_index] for row in rows if len(row) > name_index)
    finally:
        workbook.close()


def merge_address_terms(workbook_paths: list[str | Path] | tuple[str | Path, ...]) -> list[str]:
    merged: list[str] = []
    for workbook_path in workbook_paths:
        merged.extend(load_address_terms(workbook_path))
    return list(dict.fromkeys(merged))


def load_slot_terms(slots_path: str | Path) -> list[str]:
    lines = Path(slots_path).read_text(encoding="utf-8").splitlines()
    return _unique_terms(line for line in lines if not line.strip().startswith("#"))


def load_dynamic_terms(dynamic_dir: str | Path) -> list[str]:
    terms: list[str] = []
    dynamic_dir = Path(dynamic_dir)
    for relative_path in DYNAMIC_RELATIVE_FILES:
        path = dynamic_dir / relative_path
        for raw_line in path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.rsplit(" ", 1)
            value = parts[0] if len(parts) == 2 and parts[1].strip().isdigit() else line
            if term := normalize_hotword(value):
                terms.append(term)
    return list(dict.fromkeys(terms))


def build_full_hotwords(
    address_terms: list[str],
    slot_terms: list[str],
    dynamic_terms: list[str] | None = None,
) -> list[str]:
    return list(dict.fromkeys([*address_terms, *slot_terms, *(dynamic_terms or [])]))


def render_hotword_file(
    address_terms: list[str],
    slot_terms: list[str],
    dynamic_terms: list[str],
    *,
    workbook_paths: list[str | Path] | tuple[str | Path, ...],
) -> str:
    address_set = set(address_terms)
    unique_slots = [term for term in slot_terms if term not in address_set]
    address_and_slots = address_set | set(unique_slots)
    unique_dynamic = [term for term in dynamic_terms if term not in address_and_slots]
    lines = [
        "# 全量 ASR 前处理热词；由 generate_full_hotwords.py 生成，请勿手工编辑。",
        *(f"# 地址来源: {Path(path)}" for path in workbook_paths),
        f"# 地址热词: {len(address_terms)}",
        f"# 槽位热词: {len(unique_slots)}",
        f"# 动态词表补充热词: {len(unique_dynamic)}",
        f"# 总热词: {len(address_terms) + len(unique_slots) + len(unique_dynamic)}",
        "# 地址热词",
        *address_terms,
        "# 槽位热词",
        *unique_slots,
        "# 动态词表补充热词",
        *unique_dynamic,
        "",
    ]
    return "\n".join(lines)


def render_terms(terms: list[str], *, title: str) -> str:
    return "\n".join([f"# {title}", *terms, ""])


def generate(
    workbook_paths: list[Path] | tuple[Path, ...],
    slots_path: Path,
    dynamic_dir: Path,
    address_output_path: Path,
    full_output_path: Path,
) -> tuple[int, int, int]:
    address_terms = merge_address_terms(workbook_paths)
    slot_terms = load_slot_terms(slots_path)
    dynamic_terms = load_dynamic_terms(dynamic_dir)
    full_terms = build_full_hotwords(address_terms, slot_terms, dynamic_terms)
    address_output_path.parent.mkdir(parents=True, exist_ok=True)
    address_output_path.write_text(
        render_terms(address_terms, title="深圳地址全量热词"),
        encoding="utf-8",
    )
    full_output_path.write_text(
        render_hotword_file(
            address_terms,
            slot_terms,
            dynamic_terms,
            workbook_paths=workbook_paths,
        ),
        encoding="utf-8",
    )
    return len(address_terms), len(slot_terms), len(full_terms)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        dest="workbooks",
        action="append",
        type=Path,
        help="地址 Excel；可重复传入。未传时使用两个默认深圳地址表。",
    )
    parser.add_argument("--slots", type=Path, default=DEFAULT_SLOTS_FILE)
    parser.add_argument("--dynamic-dir", type=Path, default=DEFAULT_DYNAMIC_DIR)
    parser.add_argument("--address-output", type=Path, default=DEFAULT_ADDRESS_FILE)
    parser.add_argument("--full-output", type=Path, default=DEFAULT_FULL_FILE)
    args = parser.parse_args()
    workbook_paths = args.workbooks or list(DEFAULT_WORKBOOKS)

    address_count, slot_count, total_count = generate(
        workbook_paths,
        args.slots,
        args.dynamic_dir,
        args.address_output,
        args.full_output,
    )
    print(
        f"generated {total_count} hotwords: addresses={address_count}, "
        f"slots={slot_count}, output={args.full_output}"
    )


if __name__ == "__main__":
    main()
