from pathlib import Path
import sys

from openpyxl import Workbook


ASR_DIR = Path(__file__).resolve().parents[1]
if str(ASR_DIR) not in sys.path:
    sys.path.insert(0, str(ASR_DIR))

from generate_full_hotwords import (  # noqa: E402
    build_full_hotwords,
    load_address_terms,
    load_dynamic_terms,
    load_slot_terms,
    merge_address_terms,
    normalize_hotword,
)


def test_normalize_hotword_removes_brackets_but_keeps_inner_text():
    assert normalize_hotword(" 是否存在外籍客人（语言不通） ") == "是否存在外籍客人语言不通"
    assert normalize_hotword("M Stand(科兴科学园店)") == "MStand科兴科学园店"
    assert normalize_hotword("管理中心【8栋】") == "管理中心8栋"


def test_load_address_terms_uses_name_column_and_dedupes_after_normalization(tmp_path):
    workbook_path = tmp_path / "addresses.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["id", "name", "level"])
    sheet.append([1, "管理中心（8栋）", "POI-1"])
    sheet.append([2, "管理中心(8栋)", "POI-1"])
    sheet.append([3, "科技中二路", "LOI"])
    workbook.save(workbook_path)

    assert load_address_terms(workbook_path) == ["管理中心8栋", "科技中二路"]


def test_build_full_hotwords_keeps_address_order_then_adds_unique_slots(tmp_path):
    workbook_path = tmp_path / "addresses.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name"])
    sheet.append(["灾害地址"])
    sheet.append(["国人通信大厦(北门)"])
    workbook.save(workbook_path)
    slots_path = tmp_path / "slots.txt"
    slots_path.write_text("灾害地址\n灾害类型\n", encoding="utf-8")

    address_terms = load_address_terms(workbook_path)
    slot_terms = load_slot_terms(slots_path)
    terms = build_full_hotwords(address_terms, slot_terms)

    assert terms == ["灾害地址", "国人通信大厦北门", "灾害类型"]


def test_merge_address_terms_keeps_source_order_and_dedupes_across_workbooks(tmp_path):
    workbook_paths = []
    for index, names in enumerate((["深圳软件园", "东门"], ["东门", "软件园一期"])):
        workbook_path = tmp_path / f"addresses-{index}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["name"])
        for name in names:
            sheet.append([name])
        workbook.save(workbook_path)
        workbook_paths.append(workbook_path)

    assert merge_address_terms(workbook_paths) == ["深圳软件园", "东门", "软件园一期"]


def test_load_dynamic_terms_merges_all_stage_files_and_removes_weights(tmp_path):
    files = {
        "address.txt": "科兴科学园 38\n东门\n",
        "inquiry_fire_base.txt": "着火 20\n",
        "scenes/highrise.txt": "高层建筑\n",
        "scenes/crowded_place.txt": "学校\n",
        "scenes/chemical.txt": "危化品\n",
        "scenes/elevator.txt": "电梯困人\n东门\n",
    }
    for relative_path, content in files.items():
        path = tmp_path / relative_path
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(content, encoding="utf-8")

    assert load_dynamic_terms(tmp_path) == [
        "科兴科学园",
        "东门",
        "着火",
        "高层建筑",
        "学校",
        "危化品",
        "电梯困人",
    ]


def test_build_full_hotwords_adds_only_unique_dynamic_terms():
    terms = build_full_hotwords(
        ["深圳软件园", "东门"],
        ["灾害类型"],
        ["东门", "着火"],
    )

    assert terms == ["深圳软件园", "东门", "灾害类型", "着火"]
